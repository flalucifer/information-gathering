#!/usr/bin/python
# -*- coding: UTF-8 -*-
import tldextract, socket, random, hashlib, requests, time, json, execjs, threading, nmap, urllib3
from bs4 import BeautifulSoup
from queue import Queue
from openpyxl.styles import Alignment, Font, Border, Side
from libnmap.process import NmapProcess
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from global_variable import *

"""
初始化处理域名类
处理域名，拆解成：主域、主域对应ip、原始域名，原始域名对应ip
"""


class initialize:
    def __init__(self, url):
        res = tldextract.extract(url=url)
        self.tartget = res.subdomain + '.' + res.domain + '.' + res.suffix
        self.domain = res.domain + '.' + res.suffix
        self.domain_ip = socket.gethostbyname(self.domain)
        self.target_ip = socket.gethostbyname(url)


"""
ICP备案类
主要用于获取ICP备案信息
"""


class ICP:
    t_domain = ''

    def __init__(self, domain):
        ICP.t_domain = domain

    """"
    原始接口函数
    :url: 请求接口的url
    :data: 请求必要的payload参数信息
    :Content: 请求头类型，当请求是获取token时，填application/x-www-form-urlencoded;charset=UTF-8；
    当请求是正式获取备案信息时，填application/json;charset=UTF-8
    :token: 请求必要的token，当请求是获取token的时候，需要赋值"0"，否则直接带上token
    """

    @staticmethod
    def beian_post(url, data, Content, token):
        ip = "101." + str(random.randint(1, 255)) + "." + str(random.randint(1, 255)) + "." + str(
            random.randint(1, 255))
        header = {
            'Content-Type': Content,
            "Origin": "https://beian.miit.gov.cn/",
            "Referer": "https://beian.miit.gov.cn/",
            "token": token,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.87 Safari/537.36",
            'CLIENT-IP': ip,
            'X-FORWARDED-FOR': ip
        }
        r = requests.post('https://hlwicpfwc.miit.gov.cn/icpproject_query/api/' + url, data=data, headers=header).text
        return r

    """
    根据原始接口请求获取备案信息，并处理信息
    """

    @staticmethod
    def get_beian():
        print("正在收集ICP备案信息......")
        timestamp = str(int(time.time()))
        authKey = hashlib.md5(("testtest" + timestamp).encode(encoding='utf-8')).hexdigest()
        post_info = {
            "pageNum": "",
            "pageSize": "",
            "unitName": ICP.t_domain
        }
        token = json.loads(ICP.beian_post("auth", "authKey=" + authKey + "&timeStamp=" + timestamp,
                                          "application/x-www-form-urlencoded;charset=UTF-8", "0"))['params'][
            'bussiness']
        query = json.loads(
            ICP.beian_post('icpAbbreviateInfo/queryByCondition', json.dumps(post_info),
                           'application/json;charset=UTF-8',
                           token))['params']['list']
        if not token:
            icp = "服务器请求频率过高，请稍后再试"
            msg = "查询失败"
            code = "0"
            js = {
                'icp': icp,
                'msg': msg,
                'code': code
            }
        else:
            try:
                query[0]['serviceLicence']
            except:
                icp = "未备案"
                msg = "查询成功"
                code = "1"
                js = {
                    'icp': icp,
                    'msg': msg,
                    'code': code
                }
            else:
                msg = "查询成功"
                code = "1"
                icp = {
                    'ICP备案主体信息': {
                        '备案/许可证号': query[0]['mainLicence'],
                        '审核通过日期': query[0]['updateRecordTime'],
                        '主办单位名称': query[0]['unitName'],
                        '主办单位性质': query[0]['natureName']
                    },
                    'ICP备案网站信息': {
                        '网站备案/许可证号': query[0]['serviceLicence'],
                        '网站域名': query[0]['domain'],
                        '网站前置审批项': query[0]['leaderName']
                    }
                }
                js = {
                    'icp': icp,
                    'msg': msg,
                    'code': code
                }
            return js


class whois:
    def __init__(self, domain):
        self.domain = domain
        self.data = {
            'host': self.domain,
            'isUp': False,
            'ws': '',
            'token': self.get_token(self.domain)
        }

    def get_js(self):
        f = open("./js/generatetoken.js", 'r', encoding='gbk')
        line = f.readline()
        htmlstr = ''
        while line:
            htmlstr = htmlstr + line
            line = f.readline()
        return htmlstr

    def get_token(self, domain):
        js_str = self.get_js()
        ctx = execjs.compile(js_str)
        token = ctx.call('generateHostKey', domain)
        return token

    def get_info(self):
        print("正在收集whois信息......")
        whoisinfo = {}
        whoisinfo['域名'] = self.domain
        r = json.loads(requests.post(url="https://whois.chinaz.com/getWhoisInfo.ashx", data=self.data).text)['data']
        html = BeautifulSoup(r, 'html.parser')
        first_li = html.find_all("li", {"class": "clearfix"})
        # print(first_li)
        for first in first_li:
            temp = BeautifulSoup(str(first).replace('<br/>', '/'), 'lxml')
            title = (temp.find("div", {"class": "fl"}).text)
            content = (temp.find("div", {"class": "fr"}).text)
            whoisinfo[title] = content
            # print(content)
        return whoisinfo


class bp:
    def __init__(self, domain):
        self.dic = self.open_file('./dic/sub_full.txt')
        self.q = Queue()
        self.subdomain_info = {}
        self.lock=threading.Lock()
        for i in self.dic:
            self.q.put(i)
        self.pbar = tqdm(total=len(self.dic), desc='正在爆破子域名')
        threads = []
        for i in range(300):
            thread1 = threading.Thread(target=self.burp, args=(self.q, domain))
            thread1.setDaemon=True
            thread1.start()
            threads.append(thread1)

        for i in threads:
            i.join()

    def open_file(self, path):
        dic = []
        with open(path, 'r') as f:
            for i in f:
                i = i.strip()
                dic.append(i)
        return dic

    def burp(self, queue, domain):
        while True:
            if not queue.empty():
                # print('###'+(queue.qsize())+'###',end='')
                subdomain = queue.get() + '.' + domain
                self.pbar.update(1)
                try:
                    ip = socket.gethostbyname(subdomain)
                    self.lock.acquire()
                    self.subdomain_info[subdomain] = ip
                    self.lock.release()
                    # print('\033[1;32;40m %s \033[0m]\n' % (subdomain + '   ' + ip), end='')
                except Exception as e:
                    continue
            else:
                self.pbar.update(self.pbar.total-self.pbar.n)
                break



class syn_scan:
    def __init__(self, target, filename):
        self.target = target
        self.filename = filename
        self.auto_width(filename)

        self.q = Queue()
        self.IP_lists = self.scan_ip_c()
        self.lock = threading.Lock()
        self.datas = []
        for i in self.IP_lists:
            if i['状态'] == 'up':
                self.q.put(i['主机'])
        self.pbar = tqdm(total=self.q.qsize(), desc='正在扫描端口',leave=True)
        t = time.time()
        threads = []
        # 40个线程
        for i in range(70):
            thread1 = threading.Thread(target=self.scan_port, args=(self.q,))
            thread1.start()
            threads.append(thread1)
        for i in threads:
            i.join()
        time.sleep(10)
        print('端口扫描完成！')
        print("耗时：" + str(time.time() - t))


        urls = []
        for item in self.datas:
            # print(item)
            if item['port_info'] != '没有探测到端口':
                for port_info_item in item['port_info']:
                    url = ''
                    if (port_info_item['服务'] == 'https'):
                        url = "https://" + item['ip'] + ":" + port_info_item['端口']
                    elif ("http" in port_info_item['服务']):
                        url = "http://" + item['ip'] + ":" + port_info_item['端口']
                    if ("http" in url):
                        # print(url)
                        urls.append(url)
        self.pbar1 = tqdm(total=len(urls), desc='正在获取网页标题')
        if len(urls) != 0:
            threads = []
            for url in urls:
                thread1 = threading.Thread(target=self.req, args=(url,))
                thread1.start()
                threads.append(thread1)

            for thread in threads:
                thread.join()

    def scan_ip_c(self):
        prec = tqdm(total=100, position=0, desc="正在扫描C段IP存活主机",leave=True)
        res = []
        nmap_proc = NmapProcess(targets=self.target + '/24', options="-sP -PS")
        nmap_proc.run_background()
        while nmap_proc.is_running():
            time.sleep(1)
            t = (float(nmap_proc.progress) - (prec.n))
            prec.update(t)
        time.sleep(1)
        ScanEngine = nmap.PortScanner()
        r = ScanEngine.analyse_nmap_xml_scan(nmap_proc.stdout)
        for i in r['scan'].items():
            if i[1]['status']['state'] == 'up':
                res.append({'主机': i[0], '状态': i[1]['status']['state']})
        return res

    # 保存数据到excel
    def Save_Date(self, data, filename):
        # print(data)
        num = len(data['port_info'])
        filename = filename + ".xlsx"
        wb_save = load_workbook(filename)
        ws_save = wb_save['端口开放情况表']
        if data['port_info'] == '没有探测到端口':
            d = ['',data['ip'], data['os_info'], data['port_info'], '', '']
            num = 1
            ws_save.append(d)
            rows = ws_save.max_row
            # 合并单元
            ws_save.merge_cells("D" + str(rows) + ":F" + str(rows))
            wb_save.save(filename)
        else:
            for i in range(1, num + 1):
                d = ['',data['ip'], data['os_info'], data['port_info'][i - 1]['端口'], data['port_info'][i - 1]['开放状态'],
                     data['port_info'][i - 1]['服务']]
                ws_save.append(d)
        # 合并单元格并居中
        endrow_rows = ws_save.max_row
        start_rows = endrow_rows - num + 1
        # print(endrow_rows,start_rows)
        alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
        ws_save.merge_cells("A" + str(start_rows) + ":A" + str(endrow_rows))
        ws_save.merge_cells("B" + str(start_rows) + ":B" + str(endrow_rows))
        ws_save.merge_cells("C" + str(start_rows) + ":C" + str(endrow_rows))
        ws_area = ws_save["A" + str(start_rows) + ":F" + str(endrow_rows)]
        for i in ws_area:
            for j in i:
                j.alignment = alignment
                j.border = border
        wb_save.save(filename)

    # 调整单元格宽度
    def auto_width(self, filename):
        filename = filename + ".xlsx"
        wb_save = load_workbook(filename)
        ws_save = wb_save['端口开放情况表']
        ws_save.column_dimensions['A'].width = 50
        ws_save.column_dimensions['B'].width = 16.44
        ws_save.column_dimensions['C'].width = 50.44
        ws_save.column_dimensions['D'].width = 5.78
        ws_save.column_dimensions['E'].width = 8.78
        ws_save.column_dimensions['F'].width = 16.44
        wb_save.save(filename)

    # 扫描端口信息
    def scan_port(self, queue):
        while True:
            if not queue.empty():
                target = queue.get()
                nmap_proc = NmapProcess(targets=target, options="-sS -O")
                nmap_proc.run_background()
                while nmap_proc.is_running():
                    time.sleep(2)
                ScanEngine = nmap.PortScanner()
                r = ScanEngine.analyse_nmap_xml_scan(nmap_proc.stdout)
                host_port_and_os_info = {}
                port_info = []
                os_info = ""
                try:
                    tcp = list(r['scan'].values())[0]['tcp']
                    for i in tcp.items():
                        # port_info.append({'端口': str(i[0]), '开放状态': i[1]['state'], '服务': i[1]['name']})
                        if i[1]['state'] == 'open':
                            port_info.append({'端口': str(i[0]), '开放状态': i[1]['state'], '服务': i[1]['name']})
                    if port_info == []:
                        port_info = '没有探测到端口'
                except:
                    port_info = '没有探测到端口'
                    pass
                try:
                    osmatch = list(r['scan'].values())[0]['osmatch']
                    os = osmatch[0]['name']
                    os_info = os
                except:
                    os_info = "没有发现操作系统"
                host_port_and_os_info['ip'] = target
                host_port_and_os_info['port_info'] = port_info
                host_port_and_os_info['os_info'] = os_info
                # print(host_port_and_os_info)
                self.datas.append(host_port_and_os_info)
                self.lock.acquire()
                self.Save_Date(host_port_and_os_info, self.filename)
                self.pbar.update(1)
                self.lock.release()
            else:
                break

    # 请求网站标题
    def req(self, url):
        heads = {
            'User-Agent': 'Mozilla/5.0(WindowsNT10.0;Win64;x64)AppleWebKit/537.36(KHTML,likeGecko)Chrome/92.0.4515.107Safari/537.36',
            'Accept-Encoding': 'gzip,deflate'
        }
        data = {
            'url': '',
            'title': ''
        }
        try:
            urllib3.disable_warnings()
            response = requests.get(url=url, headers=heads, verify=False, timeout=7)  # 请求漏洞的url
            if response.status_code == 200:
                bs = BeautifulSoup(response.content, "html.parser")
                title = bs.find("title").text
                data['url'] = url
                if title == "":
                    data['title'] = "未匹配到标题"
                else:
                    data['title'] = title
                self.lock.acquire()
                self.save_title_data(data, self.filename)
                self.lock.release()
            else:
                pass
        except Exception as e:
            pass
        self.lock.acquire()
        self.pbar1.update(1)
        self.lock.release()

    def save_title_data(self, data, filename):
        filename = filename + ".xlsx"
        wb = load_workbook(filename)
        ws = wb['网站标题表']
        d = [data['url'], data['title']]
        ws.append(d)
        wb.save(filename)
        endrow_rows = ws.max_row
        alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
        ws.cell(row=1, column=1).alignment = alignment
        ws.cell(row=1, column=1).border = border
        ws.cell(row=1, column=2).alignment = alignment
        ws.cell(row=1, column=2).border = border
        ws_area = ws["A" + str(endrow_rows) + ":B" + str(endrow_rows)]
        for i in ws_area:
            for j in i:
                j.alignment = alignment
                j.border = border
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 25
        wb.save(filename)

