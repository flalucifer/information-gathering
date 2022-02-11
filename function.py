#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
from global_variable import *
import re

"""
初始化表格文件，生成一个excel电子表格。
:param filename: 生成电子表格的文件名
"""


def init_excel(filename):
    wb = Workbook()
    filename = filename + ".xlsx"
    # 初始化备案信息表
    ws = wb.create_sheet(index=0)
    ws.title = "ICP备案信息表"
    ws.merge_cells("A1:D1")
    ws.merge_cells("E1:G1")
    content = ['备案/许可证号', '审核通过日期', '主办单位名称', '主办单位性质', '网站备案/许可证号', '网站域名', '网站前置审批项']
    for i in range(0, 7):
        ws.cell(2, i + 1).value = content[i]
        ws.cell(2, i + 1).border = border
        ws.cell(2, i + 1).alignment = alignment
        ws.cell(3, i + 1).border = border
        ws.cell(3, i + 1).alignment = alignment
    ws["A1"].border = border
    ws["A1"].alignment = alignment
    ws["A1"].value = "ICP备案主体信息"
    ws["G1"].border = border
    ws["E1"].border = border
    ws["E1"].alignment = alignment
    ws["E1"].value = "ICP备案网站信息"
    widths = [22.22, 22.22, 35, 17, 22.22, 22.22, 20]
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for i in range(len(columns)):
        ws.column_dimensions[columns[i]].width = widths[i]
    wb.save(filename)
    # 初始化whois信息表
    ws = wb.create_sheet(index=1)
    ws.title = "whois信息表"
    ws.merge_cells("A1:B1")
    ws.cell(1, 1).value = 'whois信息'
    ws.cell(1, 1).border = border
    ws.cell(1, 1).alignment = alignment
    ws.cell(1, 2).border = border
    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 61
    wb.save(filename)
    # 初始化表子域名信息表
    ws = wb.create_sheet(index=2)
    ws.title = "子域名信息表"
    # 初始化标题
    head=['域名','ip']
    for i in range(len(head)):
        ws.cell(1, i+1).value = head[i]
        ws.cell(1, i+1).border = border
        ws.cell(1, i+1).alignment = alignment
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25



    # 初始化c段网络表
    ws = wb.create_sheet(index=3)
    ws.title = "端口开放情况表"
    head = ['domain', 'ip', 'os_info', 'port', 'state', 'protocol']
    for i in range(0, 6):
        ws.cell(1, i + 1).value = head[i]
        ws.cell(1, i + 1).border = border
        ws.cell(1, i + 1).alignment = alignment
        wb.save(filename)
    ws1 = wb.create_sheet(index=4)
    ws1.title = "网站标题表"
    head = ['url', 'title']
    for i in range(0, len(head)):
        ws1.cell(1, i + 1).value = head[i]
        wb.save(filename)
    del_sheet = wb['Sheet']
    wb.remove(del_sheet)
    wb.save(filename)


def save_ICP_data(data, filename):
    filename = filename + ".xlsx"
    wb = load_workbook(filename)
    ws = wb['ICP备案信息表']
    data1 = data['icp']['ICP备案主体信息']
    data2 = data['icp']['ICP备案网站信息']
    data3 = list(dict(data1, **data2).values())
    for i in range(0, 7):
        ws.cell(3, i + 1).value = data3[i]
    wb.save(filename)


def save_whois_data(data, filename):
    filename = filename + ".xlsx"
    wb = load_workbook(filename)
    ws = wb['whois信息表']
    items = list(data.items())
    # print(items)
    # print(len(data))
    for i in range(0, len(data)):
        ws.cell(i+2, 1).value = items[i][0]
        ws.cell(i+2, 1).border = border
        ws.cell(i+2, 1).alignment = alignment
        ws.cell(i+2, 2).value = items[i][1]
        ws.cell(i+2, 2).border = border
        ws.cell(i+2, 2).alignment = alignment
    wb.save(filename)


def input_filename(filename):
    if filename==None:
        return 'scan_res'
    else:
        return filename

def deal_subdomain_info(data,domain,domain_ip):
    try:
        data[domain]
    except:
        data[domain]=domain_ip
    return data

def save_subdomain_data(data,filename):
    items=list(data.items())
    filename = filename + ".xlsx"
    wb = load_workbook(filename)
    ws = wb['子域名信息表']
    for i in range(0,len(items)):
        ws.cell(i+2,1).value=items[i][0]
        ws.cell(i+2,2).value=items[i][1]
        ws.cell(i+2,1).border=border
        ws.cell(i+2,1).alignment=alignment
        ws.cell(i+2,2).border=border
        ws.cell(i+2,2).alignment=alignment
    wb.save(filename)



def handler(signum, frame):
    exit()
    global is_exit
    is_exit = True
    print("receive a signal %d, is_exit = %d"%(signum, is_exit))

def compare_ip_domain(filename,last_subdomain_list):
    filename = filename + ".xlsx"
    wb = load_workbook(filename)
    ws = wb['端口开放情况表']
    endrow_rows=ws.max_row
    ws_area = ws["A" + str(2) + ":F" + str(endrow_rows)]
    temp_last_subdomain_list=list(last_subdomain_list)
    for row in ws_area:
        for subdomain_index in range(0,len(temp_last_subdomain_list)):
            if row[1].value==temp_last_subdomain_list[subdomain_index][1]:
                # print('索引:',subdomain_index)
                # print('剩余域名:',last_subdomain_list)
                if row[0].value==None:
                    row[0].value='/'+temp_last_subdomain_list[subdomain_index][0]
                elif '/'+temp_last_subdomain_list[subdomain_index][0] in row[0].value:
                    pass
                else:
                    row[0].value=row[0].value+'/'+temp_last_subdomain_list[subdomain_index][0]
                try:
                    a = temp_last_subdomain_list[subdomain_index]
                    print('本次需要删除的项目:',a)
                    print('删除前剩余域名:',last_subdomain_list)
                    last_subdomain_list.remove(a)
                    print('删除后剩余域名:',last_subdomain_list)
                except:
                    print('删除域名报错！')
                    print('剩余域名长度:',len(last_subdomain_list))
                    print('剩余域名:',last_subdomain_list)
                    print('本次需要删除的项目:',last_subdomain_list[subdomain_index])
                wb.save(filename)
    print('本次没有删除项目')
    return last_subdomain_list

def deal_maxsubdomain(maxsubdomain):
    if maxsubdomain==None:
        return maxsubdomain
    else:
        try:
            int(maxsubdomain)
            return maxsubdomain
        except:
            exit()

# 判断ip是否是是之前扫描结果里面出现的同一c段ip
def judge_repleate_ip(ip,filename):
    filename = filename + ".xlsx"
    wb = load_workbook(filename)
    ws = wb['端口开放情况表']
    endrow_rows = ws.max_row
    ws_area = ws["A" + str(2) + ":B" + str(endrow_rows)]
    pattern = re.compile('^(1\\d{2}|2[0-4]\\d|25[0-5]|[1-9]\\d|[1-9])\\.' + "(1\\d{2}|2[0-4]\\d|25[0-5]|[1-9]\\d|\\d)\\." + "(1\\d{2}|2[0-4]\\d|25[0-5]|[1-9]\\d|\\d)\\.")
    for row in ws_area:
        # print(type(row[1].value))
        original_ip = re.search(pattern, row[1].value).group()
        now_ip = re.search(pattern, ip).group()
        if original_ip == now_ip:
            return True
        else:
            return False

def get_last_subdomain(filename,target_ip):
    filename = filename + ".xlsx"
    wb = load_workbook(filename)
    ws = wb['子域名信息表']
    subdomains_list=[]
    endrow_rows = ws.max_row
    ws_area = ws["A" + str(2) + ":B" + str(endrow_rows)]
    for row in ws_area:
        subdomains_list.append((row[0].value,row[1].value))
    return subdomains_list
